"""
Extract text, images, and internal links from all HTML pages and output:
- 3 CSV files (Text, Images, Links), and
- 1 XLSX with 3 sheets if openpyxl is available.
Columns: Page, Location, Text, Update
"""
import csv
import os
import re
import sys
from pathlib import Path

# Project root (parent of scripts/)
ROOT = Path(__file__).resolve().parent.parent
HTML_DIR = ROOT
OUTPUT_DIR = ROOT

# HTML files to scan (root-level only, main site pages)
HTML_GLOB = "*.html"
EXCLUDE = {"test.html", "image-review-debug.html"}  # optional exclusions


def get_text_content(html):
    """Strip tags and return visible text, with simple location hints from context."""
    # Remove script/style
    html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
    text_entries = []

    # Section comments as location
    sections = list(re.finditer(r'<!--\s*(.*?)\s*-->', html))
    # Tags that contain meaningful text
    tag_pattern = re.compile(
        r'<(p|h[1-6]|span|a|li|td|th|label|button|figcaption)[^>]*>(.*?)</\1>',
        re.DOTALL | re.IGNORECASE
    )
    # Also get title
    title_m = re.search(r'<title[^>]*>(.*?)</title>', html, re.DOTALL | re.IGNORECASE)
    if title_m:
        text_entries.append(("Title", "Document title", title_m.group(1).strip()))

    pos = 0
    for m in re.finditer(tag_pattern, html):
        inner = m.group(2)
        inner = re.sub(r'<[^>]+>', '', inner)
        inner = inner.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
        inner = ' '.join(inner.split()).strip()
        if not inner or len(inner) < 2:
            continue
        tag = m.group(1).lower()
        location = tag
        if tag.startswith('h'):
            location = "Heading"
        elif tag == 'a':
            location = "Link text"
        elif tag == 'li':
            location = "List item"
        elif tag in ('td', 'th'):
            location = "Table cell"
        text_entries.append((location, tag, inner))
    return text_entries


def strip_inner(inner):
    inner = re.sub(r'<[^>]+>', '', inner)
    inner = inner.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')
    return ' '.join(inner.split()).strip()


def get_text_entries_from_html(html, page_name):
    """Return list of dicts: Page, Location, Text, Update."""
    entries = []
    clean = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<style[^>]*>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)

    # Build list of (pos, comment_text) for section names
    comment_pos = [(m.start(), (m.group(1) or "").strip()[:80]) for m in re.finditer(r'<!--\s*(.*?)\s*-->', clean)]

    def location_for_pos(pos):
        section = "Body"
        for cpos, ctext in comment_pos:
            if cpos < pos and ctext:
                section = ctext
        return section

    seen_text = set()
    for tag in ('title', 'p', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'span', 'a', 'li', 'label', 'figcaption', 'td', 'th'):
        pattern = r'<' + tag + r'[^>]*>(.*?)</' + tag + r'>'
        for m in re.finditer(pattern, clean, re.DOTALL | re.IGNORECASE):
            inner = strip_inner(m.group(1))
            if not inner or len(inner) < 2:
                continue
            key = (page_name, inner[:500])
            if key in seen_text:
                continue
            seen_text.add(key)
            loc = "Title" if tag == "title" else location_for_pos(m.start())
            entries.append({"Page": page_name, "Location": loc, "Text": inner[:32000], "Update": ""})
    return entries


def get_image_entries(html, page_name):
    """Return list of dicts: Page, Location, Text (src/alt), Update."""
    entries = []
    # <img ... src="..." ... alt="..." ...> or just src
    for m in re.finditer(r'<img[^>]+>', html):
        tag = m.group(0)
        src = re.search(r'src=["\']([^"\']+)["\']', tag)
        alt = re.search(r'alt=["\']([^"\']*)["\']', tag)
        src_val = src.group(1) if src else ""
        alt_val = alt.group(1) if alt else ""
        location = "Image"
        if 'logo' in src_val.lower() or 'logo' in alt_val.lower():
            location = "Logo"
        elif 'banner' in src_val.lower() or 'hero' in src_val.lower():
            location = "Banner"
        entries.append({
            "Page": page_name,
            "Location": location,
            "Text": f"src={src_val} | alt={alt_val}"[:32000],
            "Update": ""
        })
    return entries


def get_link_entries(html, page_name):
    """Internal links: href to .html. Columns: Page, Location, Text (link text + href), Update."""
    entries = []
    # <a href="*.html" ...>...</a>
    pattern = re.compile(r'<a\s[^>]*href=["\']([^"\']*\.html[^"\']*)["\'][^>]*>(.*?)</a>', re.DOTALL | re.IGNORECASE)
    for m in pattern.finditer(html):
        href = m.group(1).strip()
        link_text = m.group(2)
        link_text = re.sub(r'<[^>]+>', '', link_text)
        link_text = ' '.join(link_text.split()).strip()
        if href.startswith('http') and 'mythic' not in href.lower():
            continue
        location = "Nav" if not link_text else "Link"
        entries.append({
            "Page": page_name,
            "Location": location,
            "Text": f"href={href} | text={link_text}"[:32000],
            "Update": ""
        })
    return entries


def main():
    os.chdir(HTML_DIR)
    html_files = sorted(Path(".").glob(HTML_GLOB))
    html_files = [f for f in html_files if f.name not in EXCLUDE and f.is_file()]

    rows_text = []
    rows_images = []
    rows_links = []

    for path in html_files:
        page_name = path.name
        try:
            html = path.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            print(f"Skip {page_name}: {e}", file=sys.stderr)
            continue
        rows_text.extend(get_text_entries_from_html(html, page_name))
        rows_images.extend(get_image_entries(html, page_name))
        rows_links.extend(get_link_entries(html, page_name))

    # Dedupe text by (Page, Text) keeping first
    seen = set()
    unique_text = []
    for r in rows_text:
        key = (r["Page"], r["Text"][:200])
        if key in seen:
            continue
        seen.add(key)
        unique_text.append(r)

    # Write CSVs
    fieldnames = ["Page", "Location", "Text", "Update"]
    text_path = OUTPUT_DIR / "content-audit-text.csv"
    with open(text_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(unique_text)
    print(f"Wrote {text_path} ({len(unique_text)} rows)")

    images_path = OUTPUT_DIR / "content-audit-images.csv"
    with open(images_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows_images)
    print(f"Wrote {images_path} ({len(rows_images)} rows)")

    links_path = OUTPUT_DIR / "content-audit-links.csv"
    with open(links_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows_links)
    print(f"Wrote {links_path} ({len(rows_links)} rows)")

    # XLSX with 3 sheets if openpyxl available
    try:
        import openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

        ws1 = wb.create_sheet("Text", 0)
        ws1.append(fieldnames)
        for r in unique_text:
            ws1.append([r["Page"], r["Location"], r["Text"], r["Update"]])

        ws2 = wb.create_sheet("Images", 1)
        ws2.append(fieldnames)
        for r in rows_images:
            ws2.append([r["Page"], r["Location"], r["Text"], r["Update"]])

        ws3 = wb.create_sheet("Links", 2)
        ws3.append(fieldnames)
        for r in rows_links:
            ws3.append([r["Page"], r["Location"], r["Text"], r["Update"]])

        xlsx_path = OUTPUT_DIR / "content-audit.xlsx"
        wb.save(xlsx_path)
        print(f"Wrote {xlsx_path} (3 sheets)")
    except ImportError:
        print("Install openpyxl for a single .xlsx file: pip install openpyxl", file=sys.stderr)


if __name__ == "__main__":
    main()
